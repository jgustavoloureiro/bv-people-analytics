# -*- coding: utf-8 -*-
"""Gera o xlsx de base bruta (camada Bronze), com formatação profissional."""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

BASE = "/home/claude/bv_people_analytics"

df_colab = pd.read_csv(f"{BASE}/bronze/hris_colaboradores_raw.csv")
df_deslig = pd.read_csv(f"{BASE}/bronze/hris_desligamentos_raw.csv")
df_recrut = pd.read_csv(f"{BASE}/gold/recrutamento_funil.csv")
df_wf = pd.read_csv(f"{BASE}/gold/workforce_planning.csv")

HEADER_FILL = PatternFill("solid", fgColor="203AD1")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=9.5)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="203AD1")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def write_sheet(wb, name, df, title):
    ws = wb.create_sheet(name)
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")
    ws.append([])

    headers = list(df.columns)
    ws.append(headers)
    header_row = 3
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    for _, row in df.iterrows():
        ws.append(list(row))

    for r in range(header_row+1, header_row+1+len(df)):
        for c in range(1, len(headers)+1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER

    for col_idx, h in enumerate(headers, 1):
        max_len = max(len(str(h)), df[h].astype(str).str.len().max() if len(df) > 0 else 10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 3, 10), 35)

    ws.freeze_panes = f"A{header_row+1}"

    n_rows = len(df) + header_row
    n_cols = len(headers)
    table_ref = f"A{header_row}:{get_column_letter(n_cols)}{n_rows}"
    tab = Table(displayName=f"Tabela_{name.replace(' ','_').replace('-','_')}", ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)
    return ws

wb = Workbook()
wb.remove(wb.active)

write_sheet(wb, "HRIS_Colaboradores", df_colab, "Banco BV — Base Bruta de Colaboradores (HRIS) — Camada Bronze")
write_sheet(wb, "HRIS_Desligamentos", df_deslig, "Banco BV — Desligamentos (HRIS) — Camada Bronze")
write_sheet(wb, "ATS_Recrutamento", df_recrut, "Banco BV — Funil de Recrutamento (ATS) — Últimos 12 meses")
write_sheet(wb, "Workforce_Planning", df_wf, "Banco BV — Workforce Planning — Próximos 6 meses")

# Aba README
ws_info = wb.create_sheet("README", 0)
ws_info["A1"] = "Banco BV — People Analytics — Base de Dados Bruta (Bronze)"
ws_info["A1"].font = Font(name="Arial", bold=True, size=16, color="203AD1")
ws_info.merge_cells("A1:F1")

info_lines = [
    "",
    "⚠️ AVISO: Banco BV é usado aqui apenas como referência de vaga de emprego.",
    "Todos os dados deste arquivo (colaboradores, desligamentos, recrutamento) são 100% SINTÉTICOS,",
    "gerados via Python para fins de portfólio profissional. Não representam funcionários reais.",
    "",
    "ESTRUTURA DO ARQUIVO:",
    "• HRIS_Colaboradores — base bruta de colaboradores, exatamente como viria de um export de HRIS.",
    "   Contém PROPOSITALMENTE: duplicatas, nulos, outliers e inconsistências de categoria,",
    "   simulando os problemas reais que toda base de RH tem antes de tratamento.",
    "• HRIS_Desligamentos — eventos de desligamento com tipo e motivo.",
    "• ATS_Recrutamento — funil de contratação (candidatos → triagem → entrevista → oferta → contratado).",
    "• Workforce_Planning — projeção de headcount planejado vs. projetado, próximos 6 meses.",
    "",
    "PIPELINE COMPLETO:",
    "Este xlsx representa a camada BRONZE (raw) de um pipeline medallion architecture",
    "(Bronze → Silver → Gold). O tratamento de qualidade, modelagem preditiva e dashboard final",
    "estão documentados nos scripts Python do repositório (pasta /scripts).",
    "",
    "Designed by José Gustavo Loureiro Campos Silva",
    "linkedin.com/in/josegustavoloureiro",
]
for i, line in enumerate(info_lines, 3):
    ws_info.cell(row=i, column=1, value=line).font = Font(name="Arial", size=10, bold=line.startswith(("⚠️","ESTRUTURA","PIPELINE","Designed")))
ws_info.column_dimensions["A"].width = 100

wb.save(f"{BASE}/output/Banco_BV_People_Analytics_Base_Bruta.xlsx")
print("XLSX salvo:", f"{BASE}/output/Banco_BV_People_Analytics_Base_Bruta.xlsx")
print(f"Abas: {wb.sheetnames}")
print(f"Linhas: HRIS_Colaboradores={len(df_colab)}, Desligamentos={len(df_deslig)}, Recrutamento={len(df_recrut)}, Workforce={len(df_wf)}")
